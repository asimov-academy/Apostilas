from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Padrões dos objetos
font = Font(name='Calibri',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')

fill = PatternFill(fill_type=None,
                 start_color='FFFFFFFF',
                 end_color='FF000000')


border = Border(left=Side(border_style=None,
                           color='FF000000'),
                 right=Side(border_style=None,
                            color='FF000000'),
                 top=Side(border_style=None,
                          color='FF000000'),
                 bottom=Side(border_style=None,
                             color='FF000000'),
                 diagonal=Side(border_style=None,
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style=None,
                              color='FF000000'),
                 vertical=Side(border_style=None,
                               color='FF000000'),
                 horizontal=Side(border_style=None,
                                color='FF000000')
                )
# No border_style, usar um entre: 
# {‘thin’, ‘dashed’, ‘mediumDashDot’, ‘dashDotDot’, ‘hair’, 
# ‘dotted’, ‘mediumDashDotDot’, ‘medium’, ‘double’, ‘dashDot’, 
# ‘slantDashDot’, ‘thick’, ‘mediumDashed’}


alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)

number_format = 'General'

protection = Protection(locked=True,
                         hidden=False)


# Exemplo de definição
from openpyxl import Workbook, load_workbook
wb = Workbook()
ws = wb.active

a1 = ws['A1']
d4 = ws['D4']
ft = Font(color="FF0000")
a1.font = ft
d4.font = ft
