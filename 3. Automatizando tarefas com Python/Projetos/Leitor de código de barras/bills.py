from imbox import Imbox # pip install imbox / https://pypi.org/project/imbox/
from datetime import datetime, timedelta
import pandas as pd
import uuid  # gerador de nomes únicos
from barcode_reader import *
from openpyxl import Workbook


# E-mail e senha
username = 'rodrigo.tadewald@gmail.com'
password = open('passwords/token', 'r').read()

# Conexão com servidor de leitura de e-mail
host = "imap.gmail.com"
download_folder = "boletos"

mail = Imbox(host, username=username, password=password, ssl=True, ssl_context=None, starttls=False)
messages = mail.messages(date__gt=datetime.today() - timedelta(days=30), raw='has:attachment')

wb = Workbook()
ws = wb.active
r = 1
ws.cell(row=1, column=1).value = "Assunto"
ws.cell(row=1, column=2).value = "Código de barras"
ws.cell(row=1, column=3).value = "Linha digitável"
ws.cell(row=1, column=4).value = "Filename"

path = 'no_bill'
if not os.path.isdir(path):
    os.makedirs(path)

for (uid, message) in messages:   
    if len(message.attachments) > 0:
        for attach in message.attachments:
            att_file = attach.get('filename')

            if '.pdf' in att_file:
                print(message.subject, '-', att_file)
                unique_filename = str(uuid.uuid4())

                download_path = f"{download_folder}/{unique_filename}.pdf"
                date = pd.to_datetime(message.date)

                with open(download_path, "wb") as fp:
                    fp.write(attach.get('content').read())
                
                try:
                    barcode = BarcodeReader(download_path)
                    linha_dig = linha_digitavel(barcode)
                except Exception as e:
                    barcode = False

                if not barcode:
                    os.rename(download_path, os.path.join(path, unique_filename+'.pdf'))

                else:
                    # https://blog.juno.com.br/o-que-os-numeros-boleto-bancario-significam/
                    vencimento = datetime(1997, 10, 7) + timedelta(days=int(barcode[-14:-10]))
                    vencimento = vencimento.strftime('%d/%m/%Y')
                    valor = float(barcode[-10:])/100

                    r += 1
                    # Gravando no Excel
                    ws.cell(row=r, column=1).value = message.subject
                    ws.cell(row=r, column=2).value = barcode
                    ws.cell(row=r, column=3).value = linha_dig
                    ws.cell(row=r, column=4).value = unique_filename

wb.save("boletos.xlsx")
mail.logout()
